"""
Web Scraper Analytics Dashboard
---------------------------------
Automated web scraping using Playwright, data processing with Pandas,
Excel report generation, and Power BI integration.

"""

import asyncio
from datetime import datetime
from playwright.async_api import async_playwright
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import logging
import json

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class WebScraperAnalytics:
    """Main class for web scraping and analytics generation"""
    
    def __init__(self, output_dir='./output'):
        self.output_dir = output_dir
        self.data = []
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
    async def scrape_product_data(self):
        """Scrape product data using Playwright"""
        logger.info("Starting web scraping process...")
        
        async with async_playwright() as p:
            # Launch browser in headless mode
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            try:
                # Navigate to demo e-commerce site
                await page.goto('https://webscraper.io/test-sites/e-commerce/allinone', 
                              wait_until='networkidle')
                logger.info("Page loaded successfully")
                
                # Wait for products to load
                await page.wait_for_selector('.product-wrapper')
                
                # Extract product information
                products = await page.query_selector_all('.product-wrapper')
                
                for product in products:
                    try:
                        # Extract product details
                        title_elem = await product.query_selector('.title')
                        price_elem = await product.query_selector('.price')
                        description_elem = await product.query_selector('.description')
                        rating_elem = await product.query_selector('.ratings p')
                        
                        title = await title_elem.inner_text() if title_elem else 'N/A'
                        price_text = await price_elem.inner_text() if price_elem else '$0'
                        description = await description_elem.inner_text() if description_elem else 'N/A'
                        rating_text = await rating_elem.inner_text() if rating_elem else '0'
                        
                        # Clean and parse data
                        price = float(price_text.replace('$', '').strip())
                        rating = int(rating_text.split()[0]) if rating_text else 0
                        
                        # Get category from breadcrumb or default
                        category = 'Electronics'  # Default category
                        
                        product_data = {
                            'Product Name': title.strip(),
                            'Price': price,
                            'Description': description.strip(),
                            'Rating': rating,
                            'Category': category,
                            'Scraped Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        
                        self.data.append(product_data)
                        
                    except Exception as e:
                        logger.warning(f"Error parsing product: {e}")
                        continue
                
                logger.info(f"Successfully scraped {len(self.data)} products")
                
            except Exception as e:
                logger.error(f"Error during scraping: {e}")
                raise
            finally:
                await browser.close()
    
    def process_data(self):
        """Process and analyze scraped data"""
        logger.info("Processing scraped data...")
        
        if not self.data:
            logger.warning("No data to process")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(self.data)
        
        # Data cleaning
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
        df['Rating'] = pd.to_numeric(df['Rating'], errors='coerce')
        
        # Add calculated columns
        df['Price Category'] = pd.cut(df['Price'], 
                                      bins=[0, 50, 200, 500, float('inf')],
                                      labels=['Budget', 'Mid-Range', 'Premium', 'Luxury'])
        
        # Calculate statistics
        df['Avg Category Price'] = df.groupby('Category')['Price'].transform('mean')
        df['Price vs Avg'] = ((df['Price'] - df['Avg Category Price']) / df['Avg Category Price'] * 100).round(2)
        
        logger.info("Data processing completed")
        return df
    
    def create_excel_report(self, df):
        """Generate formatted Excel report with charts"""
        logger.info("Creating Excel report...")
        
        excel_file = f'{self.output_dir}/product_analysis_{self.timestamp}.xlsx'
        
        # Create Excel writer
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Write raw data
            df.to_excel(writer, sheet_name='Raw Data', index=False)
            
            # Create summary statistics
            summary = df.groupby('Category').agg({
                'Price': ['mean', 'min', 'max', 'count'],
                'Rating': 'mean'
            }).round(2)
            summary.columns = ['Avg Price', 'Min Price', 'Max Price', 'Product Count', 'Avg Rating']
            summary.to_excel(writer, sheet_name='Summary Statistics')
            
            # Price category analysis
            price_analysis = df.groupby('Price Category').agg({
                'Product Name': 'count',
                'Price': 'mean',
                'Rating': 'mean'
            }).round(2)
            price_analysis.columns = ['Count', 'Avg Price', 'Avg Rating']
            price_analysis.to_excel(writer, sheet_name='Price Analysis')
        
        # Format Excel file
        self._format_excel(excel_file)
        
        logger.info(f"Excel report saved: {excel_file}")
        return excel_file
    
    def _format_excel(self, excel_file):
        """Apply formatting to Excel file"""
        wb = load_workbook(excel_file)
        
        # Format Raw Data sheet
        ws = wb['Raw Data']
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format Summary Statistics sheet
        if 'Summary Statistics' in wb.sheetnames:
            ws_summary = wb['Summary Statistics']
            for cell in ws_summary[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        wb.save(excel_file)
    
    def create_powerbi_dataset(self, df):
        """Create Power BI ready dataset"""
        logger.info("Creating Power BI dataset...")
        
        # Create a clean dataset for Power BI
        powerbi_file = f'{self.output_dir}/powerbi_dataset_{self.timestamp}.csv'
        
        # Select and rename columns for Power BI
        powerbi_df = df[[
            'Product Name', 'Price', 'Rating', 'Category', 
            'Price Category', 'Scraped Date'
        ]].copy()
        
        # Save to CSV (Power BI friendly format)
        powerbi_df.to_csv(powerbi_file, index=False)
        
        # Also create a JSON version for API integration
        json_file = f'{self.output_dir}/data_export_{self.timestamp}.json'
        powerbi_df.to_json(json_file, orient='records', date_format='iso', indent=2)
        
        logger.info(f"Power BI dataset saved: {powerbi_file}")
        logger.info(f"JSON export saved: {json_file}")
        
        return powerbi_file, json_file
    
    def generate_report_summary(self, df):
        """Generate text summary of analysis"""
        summary = f"""
        ============================================
        WEB SCRAPING ANALYTICS REPORT
        ============================================
        Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        
        OVERVIEW:
        - Total Products Scraped: {len(df)}
        - Categories: {df['Category'].nunique()}
        - Price Range: ${df['Price'].min():.2f} - ${df['Price'].max():.2f}
        - Average Price: ${df['Price'].mean():.2f}
        - Average Rating: {df['Rating'].mean():.1f}/5
        
        TOP 5 PRODUCTS BY PRICE:
        {df.nlargest(5, 'Price')[['Product Name', 'Price', 'Rating']].to_string(index=False)}
        
        CATEGORY BREAKDOWN:
        {df.groupby('Category')['Price'].agg(['count', 'mean']).to_string()}
        
        FILES GENERATED:
        - Excel Report: product_analysis_{self.timestamp}.xlsx
        - Power BI Dataset: powerbi_dataset_{self.timestamp}.csv
        - JSON Export: data_export_{self.timestamp}.json
        ============================================
        """
        
        return summary
    
    async def run(self):
        """Execute complete workflow"""
        logger.info("Starting Web Scraper Analytics Pipeline...")
        
        try:
            # Step 1: Scrape data
            await self.scrape_product_data()
            
            if not self.data:
                logger.error("No data scraped. Exiting...")
                return
            
            # Step 2: Process data
            df = self.process_data()
            
            if df is None or df.empty:
                logger.error("Data processing failed. Exiting...")
                return
            
            # Step 3: Create Excel report
            excel_file = self.create_excel_report(df)
            
            # Step 4: Create Power BI dataset
            powerbi_file, json_file = self.create_powerbi_dataset(df)
            
            # Step 5: Generate summary
            summary = self.generate_report_summary(df)
            print(summary)
            
            # Save summary to file
            summary_file = f'{self.output_dir}/report_summary_{self.timestamp}.txt'
            with open(summary_file, 'w') as f:
                f.write(summary)
            
            logger.info("Pipeline completed successfully!")
            
        except Exception as e:
            logger.error(f"Pipeline failed: {e}")
            raise


async def main():
    """Main entry point"""
    import os
    
    # Create output directory
    output_dir = './output'
    os.makedirs(output_dir, exist_ok=True)
    
    # Initialize and run scraper
    scraper = WebScraperAnalytics(output_dir=output_dir)
    await scraper.run()


if __name__ == '__main__':
    asyncio.run(main())
